from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.db.session import get_db
from app.db.models import Order, Work, BoatProblem, Season, Shift, OrderStatus, WorkCategory, ProblemStatus, User
from app.schemas.report import SeasonReport, OrderSummary, WorkSummary, ProblemSummary
from app.api.dependencies import get_current_user

router = APIRouter(prefix="/api/reports", tags=["reports"])


@router.get("/season/{season_id}", response_model=SeasonReport)
def get_season_report(season_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    season = db.query(Season).filter(Season.id == season_id).first()
    if not season:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Season not found")

    shifts = db.query(Shift).filter(Shift.season_id == season_id).all()
    shift_ids = [s.id for s in shifts]

    if not shift_ids:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No shifts in this season")

    orders = db.query(Order).filter(Order.shift_id.in_(shift_ids)).all()
    completed_orders = [o for o in orders if o.status == OrderStatus.COMPLETED]
    total_orders_amount = sum([o.amount for o in completed_orders])
    orders_summary = OrderSummary(
        total_amount=total_orders_amount,
        total_count=len(orders),
        pending_count=len([o for o in orders if o.status == OrderStatus.PENDING]),
        completed_count=len(completed_orders)
    )
    # Breakdown ordini per categoria (solo completati)
    orders_by_category: dict[str, float] = {}
    for o in completed_orders:
        cat = o.category
        orders_by_category[cat] = orders_by_category.get(cat, 0.0) + float(o.amount)
    # Breakdown ordini per mese (YYYY-MM)
    orders_by_month: dict[str, float] = {}
    for o in orders:
        month = o.order_date.strftime('%Y-%m') if hasattr(o.order_date, 'strftime') else str(o.order_date)
        orders_by_month[month] = orders_by_month.get(month, 0.0) + float(o.amount)
    
    works = db.query(Work).filter(Work.shift_id.in_(shift_ids)).all()
    # Filtra solo lavori completati per i totali
    completed_works = [w for w in works if w.status == OrderStatus.COMPLETED]
    works_by_category = {}
    for work in completed_works:
        cat = work.category.value
        works_by_category[cat] = works_by_category.get(cat, 0) + 1
    # Breakdown lavori per mese (YYYY-MM) - solo completati
    works_by_month: dict[str, int] = {}
    for w in completed_works:
        month = w.work_date.strftime('%Y-%m') if hasattr(w.work_date, 'strftime') else str(w.work_date)
        works_by_month[month] = works_by_month.get(month, 0) + 1
    
    works_summary = WorkSummary(
        total_count=len(works),
        pending_count=len([w for w in works if w.status == OrderStatus.PENDING]),
        completed_count=len(completed_works),
        by_category=works_by_category
    )
    
    problems = db.query(BoatProblem).filter(BoatProblem.shift_id.in_(shift_ids)).all()
    problems_summary = ProblemSummary(
        total_count=len(problems),
        open_count=len([p for p in problems if p.status == ProblemStatus.OPEN]),
        closed_count=len([p for p in problems if p.status == ProblemStatus.CLOSED])
    )
    
    shifts_data = []
    for shift in shifts:
        shift_orders = [o for o in orders if o.shift_id == shift.id]
        shift_works = [w for w in works if w.shift_id == shift.id]
        shift_problems = [p for p in problems if p.shift_id == shift.id]
        
        shifts_data.append({
            "shift_number": shift.shift_number,
            "start_date": shift.start_date.isoformat(),
            "end_date": shift.end_date.isoformat(),
            "orders_count": len(shift_orders),
            "orders_amount": sum([o.amount for o in shift_orders]),
            "works_count": len(shift_works),
            "problems_count": len(shift_problems)
        })
    
    return SeasonReport(
        season_name=season.name,
        season_year=season.year,
        total_orders_amount=total_orders_amount,
        total_orders_count=len(completed_orders),  # Solo completati
        total_works_count=len(completed_works),     # Solo completati
        total_problems_count=len(problems),
        shifts_data=shifts_data,
        orders_summary=orders_summary,
        works_summary=works_summary,
        problems_summary=problems_summary,
        orders_by_category=orders_by_category,
        orders_by_month=orders_by_month,
        works_by_month=works_by_month
    )

@router.get("/shift/{shift_id}")
def get_shift_report(shift_id: int, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    shift = db.query(Shift).filter(Shift.id == shift_id).first()
    if not shift:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Shift not found")
    
    orders = db.query(Order).filter(Order.shift_id == shift_id).all()
    works = db.query(Work).filter(Work.shift_id == shift_id).all()
    problems = db.query(BoatProblem).filter(BoatProblem.shift_id == shift_id).all()
    
    return {
        "shift_number": shift.shift_number,
        "start_date": shift.start_date.isoformat(),
        "end_date": shift.end_date.isoformat(),
        "orders": [{"id": o.id, "title": o.title, "amount": o.amount, "status": o.status.value} for o in orders],
        "works": [{"id": w.id, "title": w.title, "category": w.category.value, "status": w.status.value} for w in works],
        "problems": [{"id": p.id, "boat_id": p.boat_id, "status": p.status.value} for p in problems],
        "summary": {
            "total_orders_amount": sum([o.amount for o in orders]),
            "total_orders_count": len(orders),
            "total_works_count": len(works),
            "total_problems_count": len(problems)
        }
    }

@router.get("/season/{season_id}/export-excel")
def export_season_excel(season_id: int, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    season = db.query(Season).filter(Season.id == season_id).first()
    if not season:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Season not found")
    
    shifts = db.query(Shift).filter(Shift.season_id == season_id).all()
    shift_ids = [s.id for s in shifts]
    
    if not shift_ids:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No shifts in this season")
    
    orders = db.query(Order).filter(Order.shift_id.in_(shift_ids)).all()
    works = db.query(Work).filter(Work.shift_id.in_(shift_ids)).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Acquisti"
    
    # Styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Titolo", "Importo", "Data Acquisto", "Categoria", "Appunti", "Effettuato Da", "Turno", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows
    row = 2
    for order in orders:
        user = db.query(User).filter(User.id == order.user_id).first()
        shift = db.query(Shift).filter(Shift.id == order.shift_id).first()
        
        ws.cell(row=row, column=1).value = order.title
        ws.cell(row=row, column=2).value = order.amount
        ws.cell(row=row, column=3).value = order.order_date.isoformat()
        ws.cell(row=row, column=4).value = order.category
        ws.cell(row=row, column=5).value = order.description
        ws.cell(row=row, column=6).value = user.username if user else "Unknown"
        ws.cell(row=row, column=7).value = f"Turno {shift.shift_number}"
        ws.cell(row=row, column=8).value = order.status.value
        
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = border
        
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    
    # Secondo foglio: Lavori
    ws2 = wb.create_sheet(title="Lavori")
    headers2 = ["Titolo", "Categoria", "Data Lavoro", "Stato", "Turno", "Creato il"]
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    row2 = 2
    for w in works:
        shift = db.query(Shift).filter(Shift.id == w.shift_id).first()
        ws2.cell(row=row2, column=1).value = w.title
        ws2.cell(row=row2, column=2).value = w.category.value
        ws2.cell(row=row2, column=3).value = w.work_date.isoformat()
        ws2.cell(row=row2, column=4).value = w.status.value
        ws2.cell(row=row2, column=5).value = f"Turno {shift.shift_number}"
        ws2.cell(row=row2, column=6).value = w.created_at.isoformat()
        for col in range(1, 7):
            ws2.cell(row=row2, column=col).border = border
        row2 += 1
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 15
    ws2.column_dimensions['D'].width = 12
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 20

    # Terzo foglio: Riepilogo
    # Calcola breakdown
    orders_by_category: dict[str, float] = {}
    orders_by_month: dict[str, float] = {}
    works_by_month: dict[str, int] = {}
    for o in orders:
        orders_by_category[o.category] = orders_by_category.get(o.category, 0.0) + float(o.amount)
        month = o.order_date.strftime('%Y-%m') if hasattr(o.order_date, 'strftime') else str(o.order_date)
        orders_by_month[month] = orders_by_month.get(month, 0.0) + float(o.amount)
    for w in works:
        month = w.work_date.strftime('%Y-%m') if hasattr(w.work_date, 'strftime') else str(w.work_date)
        works_by_month[month] = works_by_month.get(month, 0) + 1

    ws3 = wb.create_sheet(title="Riepilogo")
    ws3.cell(row=1, column=1).value = f"Resoconto Stagione {season.name} ({season.year})"
    ws3.cell(row=1, column=1).font = Font(bold=True, size=14)

    # Tabella 1: Spese per Categoria
    ws3.cell(row=3, column=1).value = "Spese per Categoria"
    ws3.cell(row=3, column=1).font = Font(bold=True)
    ws3.cell(row=4, column=1).value = "Categoria"
    ws3.cell(row=4, column=2).value = "Importo"
    r = 5
    for cat, amt in orders_by_category.items():
        ws3.cell(row=r, column=1).value = cat
        ws3.cell(row=r, column=2).value = amt
        r += 1

    # Tabella 2: Spese per Mese
    r += 1
    ws3.cell(row=r, column=1).value = "Spese per Mese"
    ws3.cell(row=r, column=1).font = Font(bold=True)
    r += 1
    ws3.cell(row=r, column=1).value = "Mese"
    ws3.cell(row=r, column=2).value = "Importo"
    r += 1
    for month, amt in orders_by_month.items():
        ws3.cell(row=r, column=1).value = month
        ws3.cell(row=r, column=2).value = amt
        r += 1

    # Tabella 3: Lavori per Mese
    r += 1
    ws3.cell(row=r, column=1).value = "Lavori per Mese"
    ws3.cell(row=r, column=1).font = Font(bold=True)
    r += 1
    ws3.cell(row=r, column=1).value = "Mese"
    ws3.cell(row=r, column=2).value = "# Lavori"
    r += 1
    for month, count in works_by_month.items():
        ws3.cell(row=r, column=1).value = month
        ws3.cell(row=r, column=2).value = count
        r += 1

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Return file
    filename = f"Resoconto_{season.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )