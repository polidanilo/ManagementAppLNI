from fastapi import APIRouter, Depends, HTTPException, status, Query
from sqlalchemy.orm import Session
from sqlalchemy import or_
from app.db.session import get_db
from app.db.models import Order, Shift, User
from app.schemas.order import OrderCreate, OrderUpdate, OrderResponse
from app.api.dependencies import get_current_user
from fastapi.responses import StreamingResponse
from datetime import date
from io import BytesIO
import pandas as pd
from openpyxl.styles import PatternFill, Font

router = APIRouter(prefix="/api/orders", tags=["orders"])


def get_shift_ordinal_name(shift_number: int) -> str:
    ordinals = ['Primo', 'Secondo', 'Terzo', 'Quarto', 'Quinto', 'Sesto']
    return ordinals[shift_number - 1] if 0 < shift_number <= len(ordinals) else f"Turno {shift_number}"


@router.post("/", response_model=OrderResponse, status_code=status.HTTP_201_CREATED)
def create_order(order: OrderCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    shift = db.query(Shift).filter(Shift.id == order.shift_id).first()
    if not shift:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Shift not found")

    db_order = Order(
        title=order.title,
        description=order.description,
        amount=order.amount,
        category=order.category,
        order_date=order.order_date,
        status=order.status,
        shift_id=order.shift_id,
        user_id=current_user.id,
        notes=order.notes,
        created_by=order.created_by or current_user.username,
    )
    db.add(db_order)
    db.commit()
    db.refresh(db_order)
    return db_order

def apply_filters(query, 
                  q: str | None,
                  date_from: date | None,
                  date_to: date | None,
                  category: str | None,
                  status_filter: str | None,
                  shift_id: int | None,
                  amount_min: float | None,
                  amount_max: float | None):
    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            Order.title.ilike(like),
            Order.description.ilike(like),
            Order.category.ilike(like),
            Order.notes.ilike(like),
            Order.created_by.ilike(like),
        ))
    if date_from:
        query = query.filter(Order.order_date >= date_from)
    if date_to:
        query = query.filter(Order.order_date <= date_to)
    if category:
        query = query.filter(Order.category == category)
    if status_filter:
        query = query.filter(Order.status == status_filter)
    if shift_id:
        query = query.filter(Order.shift_id == shift_id)
    if amount_min is not None:
        query = query.filter(Order.amount >= amount_min)
    if amount_max is not None:
        query = query.filter(Order.amount <= amount_max)
    return query

def apply_sort(query, sort_by: str, order: str):
    mapping = {
        "order_date": Order.order_date,
        "amount": Order.amount,
        "created_at": Order.created_at,
        "updated_at": Order.updated_at,
        "category": Order.category,
        "status": Order.status,
        "title": Order.title,
    }
    col = mapping.get(sort_by, Order.order_date)
    return query.order_by(col.desc() if order.lower() == "desc" else col.asc())

@router.get("/", response_model=list[OrderResponse])
def list_orders(
    q: str | None = Query(None),
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    category: str | None = Query(None),
    status_filter: str | None = Query(None),
    shift_id: int | None = Query(None),
    amount_min: float | None = Query(None),
    amount_max: float | None = Query(None),
    page: int = Query(0, ge=0),
    page_size: int = Query(0, ge=0),
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1),
    sort_by: str = "order_date",
    order: str = "desc",
    db: Session = Depends(get_db)
):
    query = db.query(Order)
    query = apply_filters(query, q, date_from, date_to, category, status_filter, shift_id, amount_min, amount_max)
    query = apply_sort(query, sort_by, order)

    if page and page_size:
        offset = (page - 1) * page_size
        return query.offset(offset).limit(page_size).all()
    else:
        return query.offset(skip).limit(limit).all()

@router.get("/export")
def export_orders(
    q: str | None = Query(None),
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    category: str | None = Query(None),
    status_filter: str | None = Query(None),
    shift_id: int | None = Query(None),
    shift_ids: str | None = Query(None),  # Nuovo: shift_ids separati da virgola
    amount_min: float | None = Query(None),
    amount_max: float | None = Query(None),
    sort_by: str = "id",  # Ordinamento per ID crescente
    order: str = "asc",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    query = db.query(Order).join(Shift)
    
    # Filtra solo ordini completati
    query = query.filter(Order.status == "completed")
    
    # Gestione shift_ids multipli
    if shift_ids:
        shift_id_list = [int(x.strip()) for x in shift_ids.split(',') if x.strip()]
        query = query.filter(Order.shift_id.in_(shift_id_list))
    elif shift_id:
        query = query.filter(Order.shift_id == shift_id)
    
    # Altri filtri
    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            Order.title.ilike(like),
            Order.description.ilike(like),
            Order.notes.ilike(like)
        ))
    if date_from:
        query = query.filter(Order.order_date >= date_from)
    if date_to:
        query = query.filter(Order.order_date <= date_to)
    if category:
        query = query.filter(Order.category == category)
    if amount_min is not None:
        query = query.filter(Order.amount >= amount_min)
    if amount_max is not None:
        query = query.filter(Order.amount <= amount_max)
    
    # Ordina per shift_id e poi per data (più recenti prima)
    query = query.order_by(Order.shift_id.asc(), Order.order_date.desc())
    rows = query.all()

    # Prepara dati per Excel
    data = []
    current_shift = None
    header_rows = []  # Traccia le righe header per colorarle
    
    for r in rows:
        # Aggiungi riga vuota e header turno quando cambia
        if current_shift != r.shift_id:
            # Riga vuota di separazione (tranne per il primo turno)
            if current_shift is not None:
                data.append({
                    "Turno": "",
                    "Titolo": "",
                    "Importo": "",
                    "User": "",
                    "Categoria": "",
                    "Note": "",
                    "Data": ""
                })
            
            current_shift = r.shift_id
            shift_obj = db.query(Shift).filter(Shift.id == r.shift_id).first()
            shift_name = get_shift_ordinal_name(shift_obj.shift_number) if shift_obj else f"Turno {r.shift_id}"
            
            # Riga header turno
            header_rows.append(len(data) + 2)  # +2 perché Excel parte da 1 e c'è la riga header
            data.append({
                "Turno": shift_name,
                "Titolo": "",
                "Importo": "",
                "User": "",
                "Categoria": "",
                "Note": "",
                "Data": ""
            })
        
        # Unisci descrizione e note
        desc_notes = []
        if r.description:
            desc_notes.append(r.description)
        if r.notes:
            desc_notes.append(r.notes)
        combined_desc = " | ".join(desc_notes) if desc_notes else ""
        
        # Riga ordine
        data.append({
            "Turno": "",
            "Titolo": r.title,
            "Importo": r.amount,
            "User": r.created_by,
            "Categoria": r.category,
            "Note": combined_desc,
            "Data": r.order_date.strftime("%d/%m/%Y") if r.order_date else ""
        })
    
    df = pd.DataFrame(data)
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Acquisti")
        
        # Applica stile alle righe header turno
        worksheet = writer.sheets["Acquisti"]
        emerald_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True)
        
        for row_num in header_rows:
            # Estendi la riga verde fino alla colonna M (13 colonne)
            for col in range(1, 14):  # A-M (13 colonne)
                cell = worksheet.cell(row=row_num, column=col)
                cell.fill = emerald_fill
                cell.font = white_font
        
        # Formatta colonna Importo con simbolo euro
        for row in range(2, worksheet.max_row + 1):
            importo_cell = worksheet.cell(row=row, column=3)  # Colonna C (Importo)
            if importo_cell.value and isinstance(importo_cell.value, (int, float)):
                importo_cell.value = f"€ {importo_cell.value:.2f}"
    
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="LNIspent.xlsx"'},
    )

@router.get("/{order_id}", response_model=OrderResponse)
def get_order(order_id: int, db: Session = Depends(get_db)):
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Order not found")
    return order

@router.put("/{order_id}", response_model=OrderResponse)
def update_order(
    order_id: int,
    order_update: OrderUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    print(f"🔍 Received order_update: {order_update.dict()}")
    print(f"🔍 Order ID: {order_id}")
    print(f"🔍 User ID: {current_user.id}")

    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        print(f"❌ Order {order_id} not found")
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Order not found")

    # Allow any authenticated user to update orders (removed authorization check)
    print(f"📋 Order before update: {order.title}, user: {order.user_id}")

    update_data = order_update.dict(exclude_unset=True)
    print(f"📝 Update data (exclude_unset): {update_data}")

    # 💡 PULIZIA DATI: Rimuovi valori null/vuoti per campi obbligatori del DB
    data_to_assign = {}
    for field, value in update_data.items():
        if value is None:
            print(f"⚠️ Skipping None value for field: {field}")
            continue
        if isinstance(value, str) and value.strip() == "":
            # Non assegnare stringhe vuote per campi obbligatori del DB
            print(f"⚠️ Skipping empty string for field: {field}")
            continue
        data_to_assign[field] = value

    print(f"📝 Data to assign after cleaning: {data_to_assign}")

    for field, value in data_to_assign.items():
        print(f"🔄 Setting {field} = {value}")
        setattr(order, field, value)

    print("💾 Committing changes...")
    db.commit()
    print("✅ Refreshing order...")
    db.refresh(order)
    print(f"✅ Order updated successfully: {order.title}")
    return order

@router.delete("/{order_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_order(order_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Order not found")
    
    # Allow any authenticated user to delete orders (removed authorization check)
    
    db.delete(order)
    db.commit()